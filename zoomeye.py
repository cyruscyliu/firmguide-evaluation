#!/usr/bin/python
import os
import json
import openpyxl
from openpyxl import workbook

date_path = '/Users/ding/projects/Empirical_study/eol_date.xlsx'
date_wb = openpyxl.load_workbook(date_path)
date_sheet = date_wb['Sheet1']

date_info = []
for row in date_sheet.rows:
    date_info.append(row[0].value) 

cmd = '''
curl -X GET 'https://api.zoomeye.org/host/search?query=app:{SKU}&facets=app,device,country,service,os,port' -H "Authorization: JWT eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZGVudGl0eSI6IjcwNTY5NDYxMkBxcS5jb20iLCJpYXQiOjE1ODgzMzM3NDQsIm5iZiI6MTU4ODMzMzc0NCwiZXhwIjoxNTg4Mzc2OTQ0fQ.LNUESg8Yzqy56q8a5AAYBtGRHdIrX_-RDxpdo2GLqHE"
'''

wb_overall = workbook.Workbook()
ws_overall = wb_overall.active
ws_overall.append(['SKU', 'total', 'available', 'app:count', 'device:count', 'country:count', 'service:count', 'os:count', 'port:count'])

wb_detail = workbook.Workbook()
ws_detail = wb_detail.active
ws_detail.append(['SKU', 'continent', 'country', 'city', 'isp', 'location', 'organization', 'ip', 'service', 'device', 'port', 'app', 'banner', 'timestamp'])

for sku in date_info:
    r = os.popen(cmd.format(SKU = sku)).read()
    result = json.loads(r)
    print(result)
    l = []
    l.append(sku)
    l.append(result['total'])
    l.append(result['available'])
    if result['total'] != 0:
        l.append(str(result['facets']['product']))
        l.append(str(result['facets']['device']))
        l.append(str(result['facets']['country']))
        l.append(str(result['facets']['service']))
        l.append(str(result['facets']['os']))
        l.append(str(result['facets']['port']))
        for m in result['matches']:
            d = []
            d.append(sku)
            d.append(m['geoinfo']['continent']['names']['en'])
            d.append(m['geoinfo']['country']['names']['en'])
            d.append(m['geoinfo']['city']['names']['en'])
            d.append(m['geoinfo']['isp'])
            d.append(str(m['geoinfo']['location']))
            if 'organization' in m['geoinfo'].keys():
                d.append(m['geoinfo']['organization'])
            d.append(m['ip'])
            d.append(m['portinfo']['service'])
            d.append(m['portinfo']['device'])
            d.append(m['portinfo']['port'])
            d.append(m['portinfo']['app'])
            d.append(m['portinfo']['banner'])
            d.append(m['timestamp'])
            try:
                ws_detail.append(d)
            except openpyxl.utils.exceptions.IllegalCharacterError:
                print(d)
                continue
    print(l)
    ws_overall.append(l)
    # break

wb_overall.save('/Users/ding/projects/Empirical_study/zoomeye_overall.xlsx')
wb_detail.save('/Users/ding/projects/Empirical_study/zoomeye_detail.xlsx')
